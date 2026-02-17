"""
ExportService - Service for exporting monthly work reports to CSV, XLSX, and PDF.

All export methods accept either a WorkerMonthlySummary or a CompanyMonthlySummary
and return an io.BytesIO buffer ready to be streamed as an HTTP response.

Timestamps stored in the summaries are UTC-aware; they are converted to the
requested local timezone before being written to the output files.
"""

import io
import logging
from typing import Union

import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ..models.reports import (
    CompanyMonthlySummary,
    DailyWorkSummary,
    WorkerMonthlySummary,
)

logger = logging.getLogger(__name__)

SummaryType = Union[WorkerMonthlySummary, CompanyMonthlySummary]

_MONTH_NAMES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_PDF_HEADER_BG = colors.HexColor("#BDD7EE")
_PDF_ROW_ALT_BG = colors.HexColor("#F2F2F2")


class ExportService:
    """Service for exporting reports to CSV, XLSX and PDF files."""

    # ---------------------------------------------------------------------------
    # CSV
    # ---------------------------------------------------------------------------

    async def export_monthly_csv(
        self,
        summary: SummaryType,
        timezone: str = "Europe/Madrid",
    ) -> io.BytesIO:
        """
        Export a monthly summary as a UTF-8-with-BOM CSV file.

        The separator is a semicolon (``;``) for compatibility with the Spanish
        locale default in Microsoft Excel.  Each row represents one calendar
        day for one worker.

        Args:
            summary: Either a WorkerMonthlySummary or a CompanyMonthlySummary.
            timezone: IANA timezone used for formatting timestamps.

        Returns:
            BytesIO buffer positioned at byte 0.
        """
        tz = pytz.timezone(timezone)
        rows = self._collect_daily_rows(summary)

        text_buffer = io.StringIO()
        # UTF-8 BOM so Excel auto-detects the encoding
        text_buffer.write("\ufeff")

        header = [
            "Fecha", "DNI", "Nombre", "Empresa",
            "Entrada", "Salida", "Horas Trabajadas",
            "Pausas (min)", "Horas Extra", "Modificado",
        ]
        text_buffer.write(";".join(header) + "\r\n")

        for row in rows:
            text_buffer.write(self._format_csv_row(row, tz) + "\r\n")

        buf = io.BytesIO(text_buffer.getvalue().encode("utf-8"))
        buf.seek(0)
        return buf

    # ---------------------------------------------------------------------------
    # XLSX
    # ---------------------------------------------------------------------------

    async def export_monthly_xlsx(
        self,
        summary: SummaryType,
        timezone: str = "Europe/Madrid",
    ) -> io.BytesIO:
        """
        Export a monthly summary as an XLSX workbook with two sheets.

        Sheet 1 "Resumen": one row per worker with totals.
        Sheet 2 "Detalle Diario": one row per worker per day.

        Args:
            summary: Either a WorkerMonthlySummary or a CompanyMonthlySummary.
            timezone: IANA timezone used for formatting timestamps.

        Returns:
            BytesIO buffer positioned at byte 0.
        """
        tz = pytz.timezone(timezone)
        wb = Workbook()

        self._build_summary_sheet(wb, summary)
        self._build_detail_sheet(wb, summary, tz)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ---------------------------------------------------------------------------
    # PDF
    # ---------------------------------------------------------------------------

    async def export_monthly_pdf(
        self,
        summary: SummaryType,
        timezone: str = "Europe/Madrid",
    ) -> io.BytesIO:
        """
        Export a monthly summary as a PDF document using ReportLab Platypus.

        The document contains:
        - A title and header block (company, period, generation date).
        - For a single worker: worker name and DNI.
        - For a company: a summary table of all workers.
        - A daily detail table.
        - A compliance footer referencing art. 34.9 ET and RD-Ley 8/2019.

        Args:
            summary: Either a WorkerMonthlySummary or a CompanyMonthlySummary.
            timezone: IANA timezone used for formatting timestamps.

        Returns:
            BytesIO buffer positioned at byte 0.
        """
        tz = pytz.timezone(timezone)
        buf = io.BytesIO()

        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
            title="Informe de Registro de Jornada",
        )

        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(
            "<b>Informe de Registro de Jornada</b>",
            styles["Title"],
        ))
        story.append(Spacer(1, 0.3 * cm))

        company_id, company_name, year, month = self._extract_meta(summary)
        period_str = f"{_MONTH_NAMES_ES[month]} {year}"
        generated_str = summary.generated_at.astimezone(tz).strftime("%d/%m/%Y %H:%M")

        header_text = (
            f"<b>Empresa:</b> {company_name}&nbsp;&nbsp;&nbsp;"
            f"<b>Periodo:</b> {period_str}&nbsp;&nbsp;&nbsp;"
            f"<b>Generado:</b> {generated_str}"
        )
        story.append(Paragraph(header_text, styles["Normal"]))
        story.append(Spacer(1, 0.4 * cm))

        if isinstance(summary, WorkerMonthlySummary):
            story.append(Paragraph(
                f"<b>Trabajador:</b> {summary.worker_name} &nbsp;&nbsp; "
                f"<b>DNI/NIE:</b> {summary.worker_id_number}",
                styles["Normal"],
            ))
            story.append(Spacer(1, 0.4 * cm))

        if isinstance(summary, CompanyMonthlySummary):
            story.append(Paragraph("<b>Resumen de trabajadores</b>", styles["Heading2"]))
            story.append(Spacer(1, 0.2 * cm))
            story.append(self._build_pdf_summary_table(summary))
            story.append(Spacer(1, 0.5 * cm))

        story.append(Paragraph("<b>Detalle diario</b>", styles["Heading2"]))
        story.append(Spacer(1, 0.2 * cm))
        story.append(self._build_pdf_detail_table(summary, tz))
        story.append(Spacer(1, 0.5 * cm))

        story.append(Paragraph(
            "Generado por OpenJornada. Registro conforme al art. 34.9 ET y RD-Ley 8/2019.",
            styles["Italic"],
        ))

        doc.build(story)
        buf.seek(0)
        return buf

    # ---------------------------------------------------------------------------
    # Internal helpers — data collection
    # ---------------------------------------------------------------------------

    @staticmethod
    def _collect_daily_rows(summary: SummaryType) -> list[DailyWorkSummary]:
        """Flatten summary into an ordered list of DailyWorkSummary objects."""
        if isinstance(summary, WorkerMonthlySummary):
            return sorted(summary.daily_details, key=lambda d: d.date)

        rows: list[DailyWorkSummary] = []
        for worker in summary.workers:
            rows.extend(worker.daily_details)
        rows.sort(key=lambda d: (d.date, d.worker_name))
        return rows

    @staticmethod
    def _extract_meta(summary: SummaryType) -> tuple[str, str, int, int]:
        """Return (company_id, company_name, year, month) from either summary type."""
        return summary.company_id, summary.company_name, summary.year, summary.month

    # ---------------------------------------------------------------------------
    # Internal helpers — CSV formatting
    # ---------------------------------------------------------------------------

    @staticmethod
    def _format_csv_row(day: DailyWorkSummary, tz: pytz.BaseTzInfo) -> str:
        """Render a single DailyWorkSummary as a semicolon-delimited CSV row."""
        date_str = day.date.strftime("%d/%m/%Y")

        entry_str = (
            day.first_entry.astimezone(tz).strftime("%H:%M")
            if day.first_entry else ""
        )
        exit_str = (
            day.last_exit.astimezone(tz).strftime("%H:%M")
            if day.last_exit else ""
        )

        worked_hours = round(day.total_worked_minutes / 60, 2)
        daily_expected = 480.0
        overtime_hours = max(0.0, round((day.total_worked_minutes - daily_expected) / 60, 2))
        modified_str = "Si" if day.is_modified else "No"

        fields = [
            date_str,
            day.worker_id_number,
            day.worker_name,
            day.company_name,
            entry_str,
            exit_str,
            f"{worked_hours:.2f}",
            f"{day.total_pause_minutes:.0f}",
            f"{overtime_hours:.2f}",
            modified_str,
        ]
        return ";".join(fields)

    # ---------------------------------------------------------------------------
    # Internal helpers — XLSX building
    # ---------------------------------------------------------------------------

    def _build_summary_sheet(self, wb: Workbook, summary: SummaryType) -> None:
        """Populate the 'Resumen' sheet with per-worker monthly totals."""
        ws = wb.active
        ws.title = "Resumen"

        headers = [
            "Trabajador", "DNI", "Dias trabajados",
            "Horas totales", "Horas extra", "Estado firma",
        ]
        ws.append(headers)
        self._style_header_row(ws, 1, len(headers))

        workers = (
            [summary]
            if isinstance(summary, WorkerMonthlySummary)
            else summary.workers
        )

        for w in workers:
            overtime_hours = round(w.total_overtime_minutes / 60, 2)
            ws.append([
                w.worker_name,
                w.worker_id_number,
                w.total_days_worked,
                round(w.total_worked_minutes / 60, 2),
                overtime_hours,
                w.signature_status,
            ])

        self._auto_column_widths(ws)

    def _build_detail_sheet(
        self, wb: Workbook, summary: SummaryType, tz: pytz.BaseTzInfo
    ) -> None:
        """Populate the 'Detalle Diario' sheet with one row per day per worker."""
        ws = wb.create_sheet("Detalle Diario")

        headers = [
            "Fecha", "DNI", "Nombre", "Empresa",
            "Entrada", "Salida", "Horas trabajadas",
            "Pausas (min)", "Descansos (min)", "Modificado",
        ]
        ws.append(headers)
        self._style_header_row(ws, 1, len(headers))

        for day in self._collect_daily_rows(summary):
            entry_str = (
                day.first_entry.astimezone(tz).strftime("%H:%M")
                if day.first_entry else ""
            )
            exit_str = (
                day.last_exit.astimezone(tz).strftime("%H:%M")
                if day.last_exit else ""
            )
            ws.append([
                day.date.strftime("%d/%m/%Y"),
                day.worker_id_number,
                day.worker_name,
                day.company_name,
                entry_str,
                exit_str,
                round(day.total_worked_minutes / 60, 2),
                round(day.total_pause_minutes, 0),
                round(day.total_break_minutes, 0),
                "Si" if day.is_modified else "No",
            ])

        self._auto_column_widths(ws)

    @staticmethod
    def _style_header_row(ws, row_num: int, col_count: int) -> None:
        """Apply bold, blue-tint fill, and centred alignment to a header row."""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

    @staticmethod
    def _auto_column_widths(ws) -> None:
        """Set each column width to fit its widest cell (capped at 40 chars)."""
        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col
            )
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 4, 44)

    # ---------------------------------------------------------------------------
    # Internal helpers — PDF building
    # ---------------------------------------------------------------------------

    @staticmethod
    def _pdf_table_style() -> TableStyle:
        """Return a standard TableStyle for report tables."""
        commands = [
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_HEADER_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
        ]
        return TableStyle(commands)

    @staticmethod
    def _apply_alternating_rows(table: Table, data_len: int) -> None:
        """Apply alternating background colours to data rows (rows 1..N)."""
        for i in range(1, data_len):
            if i % 2 == 0:
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, i), (-1, i), _PDF_ROW_ALT_BG),
                ]))

    def _build_pdf_summary_table(self, summary: CompanyMonthlySummary) -> Table:
        """Build the per-worker summary table for a company report."""
        header = ["Trabajador", "DNI", "Horas totales", "Horas extra", "Dias"]
        data = [header]

        for w in summary.workers:
            data.append([
                w.worker_name,
                w.worker_id_number,
                f"{w.total_worked_minutes / 60:.2f}",
                f"{w.total_overtime_minutes / 60:.2f}",
                str(w.total_days_worked),
            ])

        col_widths = [6 * cm, 3 * cm, 3 * cm, 3 * cm, 2 * cm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(self._pdf_table_style())
        self._apply_alternating_rows(table, len(data))
        return table

    def _build_pdf_detail_table(self, summary: SummaryType, tz: pytz.BaseTzInfo) -> Table:
        """Build the daily detail table for either a worker or company report."""
        if isinstance(summary, CompanyMonthlySummary):
            header = ["Fecha", "DNI", "Nombre", "Entrada", "Salida", "Horas", "Pausas", "Estado"]
        else:
            header = ["Fecha", "Entrada", "Salida", "Horas", "Pausas (min)", "Descansos (min)", "Estado"]

        data = [header]

        for day in self._collect_daily_rows(summary):
            entry_str = (
                day.first_entry.astimezone(tz).strftime("%H:%M")
                if day.first_entry else "-"
            )
            exit_str = (
                day.last_exit.astimezone(tz).strftime("%H:%M")
                if day.last_exit else "-"
            )
            status_str = "Abierto" if day.has_open_session else ("Mod." if day.is_modified else "OK")

            if isinstance(summary, CompanyMonthlySummary):
                data.append([
                    day.date.strftime("%d/%m/%Y"),
                    day.worker_id_number,
                    day.worker_name,
                    entry_str,
                    exit_str,
                    f"{day.total_worked_minutes / 60:.2f}",
                    f"{day.total_pause_minutes:.0f}",
                    status_str,
                ])
            else:
                data.append([
                    day.date.strftime("%d/%m/%Y"),
                    entry_str,
                    exit_str,
                    f"{day.total_worked_minutes / 60:.2f}",
                    f"{day.total_pause_minutes:.0f}",
                    f"{day.total_break_minutes:.0f}",
                    status_str,
                ])

        if isinstance(summary, CompanyMonthlySummary):
            col_widths = [2.8 * cm, 2.8 * cm, 5 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm, 2.2 * cm, 2 * cm]
        else:
            col_widths = [3 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm, 3.5 * cm, 2.5 * cm]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(self._pdf_table_style())
        self._apply_alternating_rows(table, len(data))
        return table
